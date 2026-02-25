'''
Script that goes to a specificed URL (an editorial containing the top popular movies streaming now) and scrapes what it claims as the top movies. It is updated frequently. 
Also stores all this information in an excel file called "PopularMoviesToday.xlsx" in the same directory as this Python script. 
This file will contain movie title, rotten score, link to movie page, and poster link.
Uses 'Pandas', 'BeautifulSoup', and 'Openpyxl' to make all this possible
'''

import requests
import pandas as pd
from bs4 import BeautifulSoup
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.styles import Font

#Specify the URL
url = "https://editorial.rottentomatoes.com/guide/popular-movies/"
#List to store top movies
topMovies = []
#Name of excel file to save to
excelFileName = 'PopularMoviesToday.xlsx'

#Try attempting to access the specified URL. If not successful, program stops
try:
    #Send a GET request to the url
    response = requests.get(url)
    response.raise_for_status()
    
except requests.exceptions.RequestException as e:
    print(f'An error occured: {e}')

#Parse the content using Beautifulsoup
soup = BeautifulSoup(response.content, 'html.parser')

#Get ALL movie information from all movies found on the page
movies = soup.find_all('div',class_='row countdown-item')

#Go through every movie and extract the information that I want. Will extract movie title, rotten score, link to movie page, and poster link
for movie in movies:
    
    #Grab the div where movie title, movie page, and tomato score is 
    movieTitleDiv = movie.find(class_='meta-data-wrapper')
    
    #Get the movie title
    movieTitle = movieTitleDiv.find('a').getText(strip=True)
    
    #Get the movie info link
    movieInfoLink = movieTitleDiv.find('a').get('href')
    
    #Get the tomato score
    movieTomatoScore = movieTitleDiv.find(class_='meta-scores-wrapper').get_text(strip=True)
    if (movieTomatoScore == '- -'):
        movieTomatoScore = 0.0
    else:
        movieTomatoScore = int(movieTomatoScore.rstrip('%')) / 100
    
    #For fun: grab the poster link
    moviePosterLink = movie.find(class_='article_poster').get('src')
    
    #Store all this information for use later
    aMovie = {'Movie Title': movieTitle, 'Tomato Rating': movieTomatoScore, 'Information Page': movieInfoLink, 'Poster': moviePosterLink}
    topMovies.append(aMovie)


#Convert into a dataframe
df = pd.DataFrame(topMovies)
#Create the excel file, with index set as False so rows do not have random numbers
df.to_excel(excelFileName, index=False)

'''
Now to make this workbook look good. 
'''
wb = load_workbook(filename=excelFileName)
ws = wb.active

#Make the columns spaced out nicely
for column in ws.columns:
    
    currentColumnLetter = get_column_letter(column[0].col_idx)
    
    #For every cell in the column, calculate the max width they should be
    maxWidth = 0
    for cell in column:
        #If there is something in the cell, calculate the new max width
        if (cell.value):
            maxWidth = max(maxWidth, len(str(cell.value)))
            
    #Adjust the column dimensions for the current column with an offset of 2
    ws.column_dimensions[currentColumnLetter].width = maxWidth + 2

#Test: just modifying the column titles
for cell in ws[1]:
    cell.font = Font(italic=True, bold=True)
    cell.value = cell.value.capitalize()

#Shorten the poster links. They are mega long.
for row in range(2, ws.max_row + 1):
    cell = ws[f'D{row}']
    posterLink = cell.value
    cell.value = f'=HYPERLINK("{posterLink}","Link to poster!")'
    cell.style = "Hyperlink"
#Resize again since it's not gigantic anymore
ws.column_dimensions['D'].width = 15

#Third thing, instead of the full link to the movie page, shorten it as well 
for (cell,) in ws[f'C2:C{ws.max_row}']:
    url = cell.value
    cell.value = f'=HYPERLINK("{url}","Link to Rotten page!")'
    cell.style = "Hyperlink"
#Resize since it's not gigantic anymore
ws.column_dimensions['C'].width = 20

#Lastly, as for data, make the percentage columns into numbers
for row in range(2, ws.max_row + 1):
    cell = ws[f'B{row}']
    cell.number_format = '0%'
    cell.font = Font(bold=True)
    cell.alignment = Alignment(horizontal='center', vertical='center')


'''
Adding style to the table
'''
tableCells = f'A1:D{ws.max_row}'
#Create a table
theTable = Table(displayName='TopMoviesTable', ref=tableCells)
#Create a style
style = TableStyleInfo(name='TableStyleMedium17',showColumnStripes=True, showFirstColumn=True)
#Apply the style
theTable.tableStyleInfo = style

#Save changes to the workshee, including adding the table
ws.add_table(theTable)
wb.save(excelFileName)

print("Successfully scraped information from: " + url + " into file: " + excelFileName)

    

